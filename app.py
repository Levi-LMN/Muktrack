from flask import Flask, render_template, request, redirect, url_for, jsonify, session, flash
from datetime import datetime
from collections import defaultdict
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from sqlalchemy import func
from sqlalchemy import func
from dateutil.relativedelta import relativedelta

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///funds.db'
app.config['SECRET_KEY'] = 'SRTENAS74293URONSHDGCVYWZKAUQK098387412SYYVEDSA'  # Change this to a random secret key
db = SQLAlchemy(app)



class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(80), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(20), nullable=False, default='member')
    is_password_changed = db.Column(db.Boolean, default=False)
    deposits = db.relationship('Deposit', backref='user', lazy='dynamic')
    withdrawals = db.relationship('Withdrawal', backref='user', lazy='dynamic')

    def __init__(self, name, email, password, role='member'):
        self.name = name
        self.email = email
        self.password = generate_password_hash(password)
        self.role = role

    def check_password(self, password):
        return check_password_hash(self.password, password)

    def total_deposited(self):
        return sum(deposit.amount for deposit in self.deposits)

    def total_withdrawn(self):
        return sum(withdrawal.amount for withdrawal in self.withdrawals)

    def balance(self):
        return self.total_deposited() - self.total_withdrawn()

    def deposit_months(self):
        return db.session.query(
            func.strftime('%Y-%m', Deposit.date).label('month')  # Format to 'YYYY-MM'
        ).filter(Deposit.user_id == self.id).distinct().all()


class Deposit(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    amount = db.Column(db.Integer, nullable=False)
    date = db.Column(db.Date, nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)


class Withdrawal(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    amount = db.Column(db.Integer, nullable=False)
    date = db.Column(db.Date, nullable=False)
    reason = db.Column(db.String(200), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    withdrawal_type = db.Column(db.String(20), nullable=False)  # 'personal' or 'group'

class AuditLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    action = db.Column(db.String(200), nullable=False)
    timestamp = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)

    user = db.relationship('User', backref=db.backref('audit_logs', lazy='dynamic'))

def log_action(user_id, action):
    log = AuditLog(user_id=user_id, action=action)
    db.session.add(log)
    db.session.commit()

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)
    return decorated_function




def password_change_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        user_id = session.get('user_id')
        if not user_id:
            return redirect(url_for('login'))

        user = User.query.get(user_id)
        if not user:
            session.pop('user_id', None)  # Clear invalid session
            return redirect(url_for('login'))

        if not user.is_password_changed:
            return redirect(url_for('change_password'))

        return f(*args, **kwargs)

    return decorated_function

@app.route('/')
def home():
    return render_template('main.html')


@app.route('/index')
@login_required
def index():
    users = User.query.all()
    total_deposited = sum(user.total_deposited() for user in users)

    total_personal_withdrawn = sum(
        db.session.query(func.sum(Withdrawal.amount))
        .filter(Withdrawal.user_id == user.id, Withdrawal.withdrawal_type == 'personal')
        .scalar() or 0
        for user in users
    )

    total_group_withdrawn = sum(
        db.session.query(func.sum(Withdrawal.amount))
        .filter(Withdrawal.user_id == user.id, Withdrawal.withdrawal_type == 'group')
        .scalar() or 0
        for user in users
    )

    total_withdrawn = total_personal_withdrawn + total_group_withdrawn
    total_balance = total_deposited - total_withdrawn

    user_withdrawals = {
        user.id: {
            'personal': db.session.query(func.sum(Withdrawal.amount))
                .filter(Withdrawal.user_id == user.id, Withdrawal.withdrawal_type == 'personal')
                .scalar() or 0,
            'group': db.session.query(func.sum(Withdrawal.amount))
                .filter(Withdrawal.user_id == user.id, Withdrawal.withdrawal_type == 'group')
                .scalar() or 0
        } for user in users
    }

    return render_template('index.html',
                           users=users,
                           total_deposited=total_deposited,
                           total_personal_withdrawn=total_personal_withdrawn,
                           total_group_withdrawn=total_group_withdrawn,
                           total_withdrawn=total_withdrawn,
                           total_balance=total_balance,
                           user_withdrawals=user_withdrawals)

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        user = User.query.get(session['user_id'])
        if not user or user.role != 'admin':
            flash('You need to be an admin to access this page.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)

    return decorated_function


@app.route('/users', methods=['GET', 'POST'])
@login_required
@admin_required
def users():
    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']
        password = request.form['password']
        role = request.form['role']

        existing_user = User.query.filter_by(email=email).first()
        if existing_user:
            flash('Email already registered.', 'danger')
            return redirect(url_for('users'))

        new_user = User(name=name, email=email, password=password, role=role)
        db.session.add(new_user)
        db.session.commit()
        log_action(session['user_id'], f"Created new user: {name} ({email}) with role: {role}")
        flash('New user added successfully.', 'success')
        return redirect(url_for('users'))

    users = User.query.all()
    return render_template('users.html', users=users)


from sqlalchemy import func


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email'].strip().lower()  # Normalize the email to lower case
        password = request.form['password']

        # Use func.lower() for case-insensitive email query
        user = User.query.filter(func.lower(User.email) == email).first()

        if user and user.check_password(password):
            session['user_id'] = user.id
            log_action(user.id, f"User logged in: {user.email}")
            if not user.is_password_changed:
                return redirect(url_for('change_password'))
            return redirect(url_for('user_dashboard'))

        flash('Invalid email or password', 'danger')

    return render_template('login.html')


@app.route('/logout')
def logout():
    session.pop('user_id', None)
    return redirect(url_for('index'))

@app.route('/change_password', methods=['GET', 'POST'])
@login_required
def change_password():
    user = User.query.get(session['user_id'])
    if request.method == 'POST':
        new_password = request.form['new_password']
        user.password = generate_password_hash(new_password)
        user.is_password_changed = True
        db.session.commit()
        log_action(user.id, "Changed password")
        return redirect(url_for('user_dashboard'))
    return render_template('change_password.html')

@app.route('/deposit/<int:user_id>', methods=['GET', 'POST'])
@login_required
@password_change_required
def deposit(user_id):
    user = User.query.get(user_id)
    if request.method == 'POST':
        amount = int(request.form['amount'])
        date = datetime.strptime(request.form['date'], '%Y-%m-%d').date()
        deposit = Deposit(amount=amount, date=date, user=user)
        db.session.add(deposit)
        db.session.commit()
        log_action(user.id, f"Deposited {amount} on {date}")
        return redirect(url_for('user_dashboard'))
    return render_template('deposit.html', user=user)


# Update the personal_withdrawal route

@app.route('/personal_withdrawal/<int:user_id>', methods=['GET', 'POST'])
@login_required
@password_change_required
def personal_withdrawal(user_id):
    user = User.query.get_or_404(user_id)
    current_user = User.query.get(session['user_id'])

    if current_user.id != user.id and current_user.role != 'admin':
        flash('You can only make personal withdrawals from your own account.', 'danger')
        return redirect(url_for('user_dashboard'))

    if request.method == 'POST':
        amount = int(request.form['amount'])
        date = datetime.strptime(request.form['date'], '%Y-%m-%d').date()
        reason = request.form['reason']

        if amount > user.balance():
            flash('Insufficient funds for this withdrawal.', 'danger')
            return redirect(url_for('personal_withdrawal', user_id=user.id))

        withdrawal = Withdrawal(amount=amount, date=date, reason=reason, user=user, withdrawal_type='personal')
        db.session.add(withdrawal)

        action_description = f"Personal withdrawal of {amount} on {date} for reason: {reason}"
        if current_user.id != user.id:
            action_description = f"Admin {current_user.name} initiated: " + action_description

        log_action(user.id, action_description)

        db.session.commit()

        flash('Personal withdrawal successful.', 'success')
        return redirect(url_for('user_dashboard'))

    return render_template('personal_withdrawal.html', user=user, current_user=current_user)


#Group withdrawal
from datetime import datetime
from sqlalchemy import func
import logging

# Set up logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

@app.route('/withdraw', methods=['GET', 'POST'])
@login_required
@password_change_required
def withdraw():
    if request.method == 'POST':
        withdrawal_date = datetime.strptime(request.form['date'], '%Y-%m-%d').date()
        reason = "Group withdrawal - " + request.form['reason']
        total_amount = int(request.form['amount'])

        logger.info(f"Initiating group withdrawal of {total_amount} on {withdrawal_date}")

        # Get all users who have made deposits up to the withdrawal date
        eligible_users = (db.session.query(User)
                          .join(Deposit)
                          .filter(Deposit.date <= withdrawal_date)
                          .group_by(User.id)
                          .having(func.sum(Deposit.amount) > 0)
                          .all())

        if not eligible_users:
            logger.warning("No eligible users found for withdrawal")
            flash('No eligible users found for withdrawal', 'danger')
            return redirect(url_for('withdraw'))

        # Calculate user balances and total eligible balance
        user_balances = {}
        total_eligible_balance = 0
        total_deposits = 0
        total_withdrawals = 0

        for user in eligible_users:
            deposits = db.session.query(func.sum(Deposit.amount)).filter(
                Deposit.user_id == user.id,
                Deposit.date <= withdrawal_date
            ).scalar() or 0

            withdrawals = db.session.query(func.sum(Withdrawal.amount)).filter(
                Withdrawal.user_id == user.id,
                Withdrawal.date <= withdrawal_date
            ).scalar() or 0

            user_balance = deposits - withdrawals

            logger.debug(f"User {user.id}: Deposits = {deposits}, Withdrawals = {withdrawals}, Balance = {user_balance}")

            if user_balance > 0:
                user_balances[user.id] = user_balance
                total_eligible_balance += user_balance
                total_deposits += deposits
                total_withdrawals += withdrawals

        logger.info(f"Total deposits: {total_deposits}")
        logger.info(f"Total withdrawals: {total_withdrawals}")
        logger.info(f"Calculated total eligible balance: {total_eligible_balance}")

        if total_eligible_balance < total_deposits - total_withdrawals:
            logger.error(f"Discrepancy detected: Eligible balance ({total_eligible_balance}) is less than expected ({total_deposits - total_withdrawals})")
            flash(f'Error: Calculated balance ({total_eligible_balance}) is less than expected ({total_deposits - total_withdrawals}). Please contact the administrator.', 'danger')
            return redirect(url_for('withdraw'))

        if total_eligible_balance < total_amount:
            logger.warning(f"Withdrawal amount ({total_amount}) exceeds total eligible balance ({total_eligible_balance})")
            flash(f'Total withdrawal amount ({total_amount}) exceeds total eligible balance ({total_eligible_balance})', 'danger')
            return redirect(url_for('withdraw'))

        # Calculate initial proportional contributions
        user_withdrawals = {user_id: (balance / total_eligible_balance) * total_amount
                            for user_id, balance in user_balances.items()}

        # Adjust for users who would contribute more than their balance
        excess = 0
        users_at_max = set()
        for user_id, withdrawal in user_withdrawals.items():
            if withdrawal > user_balances[user_id]:
                excess += withdrawal - user_balances[user_id]
                user_withdrawals[user_id] = user_balances[user_id]
                users_at_max.add(user_id)

        # Redistribute excess among remaining users
        while excess > 0.01 and len(users_at_max) < len(user_balances):
            remaining_users = [uid for uid in user_balances if uid not in users_at_max]
            share = excess / len(remaining_users)
            new_excess = 0
            for user_id in remaining_users:
                if user_withdrawals[user_id] + share > user_balances[user_id]:
                    new_excess += user_withdrawals[user_id] + share - user_balances[user_id]
                    user_withdrawals[user_id] = user_balances[user_id]
                    users_at_max.add(user_id)
                else:
                    user_withdrawals[user_id] += share
            excess = new_excess

        # Verify total withdrawal matches requested amount
        total_withdrawal = sum(user_withdrawals.values())
        if abs(total_withdrawal - total_amount) > 0.01:
            logger.error(f"Calculation error: Total withdrawal ({total_withdrawal}) does not match requested amount ({total_amount})")
            flash('Error in withdrawal calculation. Please contact the administrator.', 'danger')
            return redirect(url_for('withdraw'))

        # Create withdrawal records
        try:
            for user_id, withdrawal_amount in user_withdrawals.items():
                if withdrawal_amount > 0:
                    user = User.query.get(user_id)
                    withdrawal = Withdrawal(
                        amount=round(withdrawal_amount, 2),
                        date=withdrawal_date,
                        reason=reason,
                        user=user,
                        withdrawal_type='group'
                    )
                    db.session.add(withdrawal)
                    log_action(user.id, f"Group withdrawal of {withdrawal_amount:.2f} on {withdrawal_date} for reason: {reason}")

            db.session.commit()
            logger.info("Group withdrawal completed successfully")
            flash('Group withdrawal completed successfully', 'success')
        except Exception as e:
            db.session.rollback()
            logger.error(f"Error during withdrawal creation: {str(e)}")
            flash('An error occurred while processing the withdrawal. Please try again.', 'danger')

        return redirect(url_for('index'))

    return render_template('withdrawal.html')

# Update the user_dashboard route
@app.route('/user_dashboard')
@login_required
@password_change_required
def user_dashboard():
    user = User.query.get(session['user_id'])

    page_deposit = request.args.get('page_deposit', 1, type=int)
    page_personal_withdrawal = request.args.get('page_personal_withdrawal', 1, type=int)
    page_group_withdrawal = request.args.get('page_group_withdrawal', 1, type=int)

    deposits = user.deposits.order_by(Deposit.date.desc()).paginate(page=page_deposit, per_page=5, error_out=False)

    personal_withdrawals = user.withdrawals.filter_by(withdrawal_type='personal').order_by(
        Withdrawal.date.desc()).paginate(page=page_personal_withdrawal, per_page=5, error_out=False)

    group_withdrawals = user.withdrawals.filter_by(withdrawal_type='group').order_by(Withdrawal.date.desc()).paginate(
        page=page_group_withdrawal, per_page=5, error_out=False)

    return render_template('user_dashboard.html', user=user, deposits=deposits,
                           personal_withdrawals=personal_withdrawals,
                           group_withdrawals=group_withdrawals)

@app.route('/user_analytics/<int:user_id>')
@login_required
@password_change_required
def user_analytics(user_id):
    user = User.query.get_or_404(user_id)

    # Calculate deposit frequency
    first_deposit_date = db.session.query(func.min(Deposit.date)).filter(Deposit.user_id == user.id).scalar()
    if first_deposit_date:
        total_months = (
            (datetime.now().year - first_deposit_date.year) * 12 +
            (datetime.now().month - first_deposit_date.month) + 1
        )
        total_deposit_months = len(user.deposit_months())
        deposit_frequency = (total_deposit_months / total_months) * 100
    else:
        total_months = 0
        deposit_frequency = 0

    # Initialize monthly_deposits
    monthly_deposits = []

    # Generate monthly deposit history
    if first_deposit_date:
        current_date = datetime.now().date().replace(day=1)

        # Fetch all deposits at once using strftime to group by month
        deposits = Deposit.query.filter(Deposit.user_id == user.id,
                                        Deposit.date >= first_deposit_date).all()

        # Create a dictionary to group deposits by month
        monthly_deposit_data = {}
        for deposit in deposits:
            month_key = deposit.date.strftime('%Y-%m')  # Format to 'YYYY-MM'
            if month_key not in monthly_deposit_data:
                monthly_deposit_data[month_key] = {
                    'deposited': True,
                    'amount': deposit.amount
                }
            else:
                monthly_deposit_data[month_key]['amount'] += deposit.amount

        # Build the monthly deposits list
        while current_date >= first_deposit_date.replace(day=1):
            month_key = current_date.strftime('%Y-%m')
            monthly_deposits.append({
                'date': current_date,
                'deposited': monthly_deposit_data.get(month_key, {}).get('deposited', False),
                'amount': monthly_deposit_data.get(month_key, {}).get('amount', 0)
            })
            current_date -= relativedelta(months=1)

        monthly_deposits.reverse()

    return render_template('user_analytics.html', user=user,
                           deposit_frequency=round(deposit_frequency, 2),
                           monthly_deposits=monthly_deposits)


@app.route('/bulk_deposit', methods=['GET', 'POST'])
@login_required
@password_change_required
def bulk_deposit():
    if request.method == 'POST':
        data = request.json
        for entry in data:
            date = datetime.strptime(entry['date'], '%Y-%m-%d').date()
            for user_id in entry['users']:
                user = User.query.get(user_id)
                if user:
                    deposit = Deposit(amount=200, date=date, user=user)
                    db.session.add(deposit)
                    log_action(user.id, f"Bulk deposit of 200 on {date}")
        db.session.commit()
        return jsonify({"message": "Deposits recorded successfully"}), 200

    users = User.query.all()
    return render_template('bulk_deposit.html', users=users)

@app.route('/get_users', methods=['GET'])
@login_required
@password_change_required
def get_users():
    users = User.query.all()
    return jsonify([{"id": user.id, "name": user.name} for user in users])

@app.route('/audit_log')
@login_required
@password_change_required
def audit_log():
    page = request.args.get('page', 1, type=int)
    per_page = 20  # Number of logs per page
    logs = AuditLog.query.order_by(AuditLog.timestamp.desc()).paginate(page=page, per_page=per_page, error_out=False)
    return render_template('audit_log.html', logs=logs)

def format_number(value):
    return "{:,.0f}".format(value)

app.jinja_env.filters['format_number'] = format_number

@app.route('/database_management')
@login_required
@password_change_required
def database_management():
    users = User.query.all()
    deposits = Deposit.query.all()
    withdrawals = Withdrawal.query.all()
    audit_logs = AuditLog.query.all()
    return render_template('database_management.html', users=users, deposits=deposits, withdrawals=withdrawals, audit_logs=audit_logs)

@app.route('/edit/<string:model>/<int:id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_entry(model, id):
    if model == 'user':
        entry = User.query.get_or_404(id)
    elif model == 'deposit':
        entry = Deposit.query.get_or_404(id)
    elif model == 'withdrawal':
        entry = Withdrawal.query.get_or_404(id)
    else:
        abort(404)

    if request.method == 'POST':
        old_data = str(entry.__dict__)
        if model == 'user':
            entry.name = request.form['name']
            entry.email = request.form['email']
            entry.role = request.form['role']
        elif model == 'deposit' or model == 'withdrawal':
            entry.amount = int(request.form['amount'])
            entry.date = datetime.strptime(request.form['date'], '%Y-%m-%d').date()
        if model == 'withdrawal':
            entry.reason = request.form['reason']

        db.session.commit()
        new_data = str(entry.__dict__)
        log_action(session['user_id'], f"Edited {model} (ID: {id}). Old data: {old_data}. New data: {new_data}")
        flash('Entry updated successfully', 'success')
        return redirect(url_for('database_management'))

    return render_template('edit_entry.html', entry=entry, model=model)


@app.route('/delete/<string:model>/<int:id>')
@login_required
@password_change_required
def delete_entry(model, id):
    if model == 'user':
        entry = User.query.get_or_404(id)
    elif model == 'deposit':
        entry = Deposit.query.get_or_404(id)
    elif model == 'withdrawal':
        entry = Withdrawal.query.get_or_404(id)
    elif model == 'auditlog':
        entry = AuditLog.query.get_or_404(id)
    else:
        abort(404)

    entry_data = str(entry.__dict__)
    db.session.delete(entry)
    db.session.commit()
    log_action(session['user_id'], f"Deleted {model} (ID: {id}). Data: {entry_data}")
    flash('Entry deleted successfully', 'success')
    return redirect(url_for('database_management'))

@app.context_processor
def inject_user():
    if 'user_id' in session:
        user = User.query.get(session['user_id'])
        return dict(current_user=user)
    return dict(current_user=None)

def format_date(date):
    day = date.day
    suffix = 'th' if 11 <= day <= 13 else {1: 'st', 2: 'nd', 3: 'rd'}.get(day % 10, 'th')
    return date.strftime(f"%d{suffix} %B %Y")

app.jinja_env.filters['format_date'] = format_date

from sqlalchemy import func

from sqlalchemy import func

@app.route('/duplicate_deposits')
@login_required
@admin_required
def duplicate_deposits():
    duplicate_deposits = db.session.query(
        Deposit.user_id,
        Deposit.date,
        Deposit.amount,
        func.count('*').label('count'),
        func.min(Deposit.id).label('min_id')
    ).group_by(
        Deposit.user_id,
        Deposit.date,
        Deposit.amount
    ).having(func.count('*') > 1).all()

    duplicate_data = []
    for dup in duplicate_deposits:
        user = User.query.get(dup.user_id)
        duplicate_data.append({
            'user_name': user.name,
            'date': dup.date,
            'amount': dup.amount,
            'count': dup.count,
            'min_id': dup.min_id
        })

    return render_template('duplicate_deposits.html', duplicate_data=duplicate_data)

from sqlalchemy import and_

@app.route('/delete_duplicates', methods=['POST'])
@login_required
@admin_required
def delete_duplicates():
    delete_ids = request.form.getlist('delete_ids')

    for min_id in delete_ids:
        min_id = int(min_id)
        base_deposit = Deposit.query.get(min_id)
        if base_deposit:
            # Find all duplicates except the one with min_id
            duplicates = Deposit.query.filter(
                and_(
                    Deposit.user_id == base_deposit.user_id,
                    Deposit.date == base_deposit.date,
                    Deposit.amount == base_deposit.amount,
                    Deposit.id != min_id
                )
            ).all()

            # Delete the duplicates
            for duplicate in duplicates:
                db.session.delete(duplicate)
                log_action(session['user_id'], f"Deleted duplicate deposit (ID: {duplicate.id}) for user {duplicate.user.name} on {duplicate.date}")

    db.session.commit()
    flash('Selected duplicate deposits have been deleted.', 'success')
    return redirect(url_for('duplicate_deposits'))


from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from flask import send_file
from sqlalchemy import extract
from calendar import month_name

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from flask import send_file
from sqlalchemy import extract
from calendar import month_name
from datetime import datetime


@app.route('/download_deposit_report')
@login_required
@admin_required
def download_deposit_report():
    deposits = Deposit.query.order_by(Deposit.date).all()
    withdrawals = Withdrawal.query.order_by(Withdrawal.date).all()
    users = User.query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    subheader_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    total_fill = PatternFill(start_color="E6B8B7", end_color="E6B8B7", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Summary sheet
    ws['A1'] = "Deposit and Withdrawal Summary Report"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal="center")

    # Deposit Analytics
    ws['A3'] = "Deposit Analytics"
    ws['A3'].font = Font(bold=True)
    ws.merge_cells('A3:D3')
    ws['A3'].fill = subheader_fill

    deposit_headers = ["Total Deposits", "Total Amount", "Average Amount", "Number of Months"]
    for col, header in enumerate(deposit_headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    total_deposits = len(deposits)
    total_deposit_amount = sum(deposit.amount for deposit in deposits)
    avg_deposit = total_deposit_amount / total_deposits if total_deposits > 0 else 0
    num_months = len(set((deposit.date.year, deposit.date.month) for deposit in deposits))

    for col, value in enumerate([total_deposits, total_deposit_amount, avg_deposit, num_months], start=1):
        cell = ws.cell(row=5, column=col, value=value)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Withdrawal Analytics
    ws['A7'] = "Withdrawal Analytics"
    ws['A7'].font = Font(bold=True)
    ws.merge_cells('A7:D7')
    ws['A7'].fill = subheader_fill

    withdrawal_headers = ["Total Withdrawals", "Personal Amount", "Group Amount", "Total Amount"]
    for col, header in enumerate(withdrawal_headers, start=1):
        cell = ws.cell(row=8, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    total_withdrawals = len(withdrawals)
    total_personal = sum(w.amount for w in withdrawals if w.withdrawal_type == 'personal')
    total_group = sum(w.amount for w in withdrawals if w.withdrawal_type == 'group')
    total_withdrawal_amount = total_personal + total_group

    for col, value in enumerate([total_withdrawals, total_personal, total_group, total_withdrawal_amount], start=1):
        cell = ws.cell(row=9, column=col, value=value)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Overall Balance
    ws['A11'] = "Overall Fund Balance"
    ws['A11'].font = Font(bold=True)
    ws.merge_cells('A11:D11')
    ws['A11'].fill = total_fill

    balance_headers = ["Total Deposits", "Total Withdrawals", "Current Balance", "Active Users"]
    for col, header in enumerate(balance_headers, start=1):
        cell = ws.cell(row=12, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    current_balance = total_deposit_amount - total_withdrawal_amount
    active_users = len([user for user in users if user.balance() > 0])

    for col, value in enumerate([total_deposit_amount, total_withdrawal_amount, current_balance, active_users],
                                start=1):
        cell = ws.cell(row=13, column=col, value=value)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        cell.font = Font(bold=True)

    # Monthly Trend
    ws['A15'] = "Monthly Deposit Trend"
    ws['A15'].font = Font(bold=True)
    ws.merge_cells('A15:D15')
    ws['A15'].fill = subheader_fill

    trend_headers = ["Month", "Number of Deposits", "Total Amount", "Average Amount"]
    for col, header in enumerate(trend_headers, start=1):
        cell = ws.cell(row=16, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    months = sorted(set((deposit.date.year, deposit.date.month) for deposit in deposits))
    for row, (year, month) in enumerate(months, start=17):
        month_deposits = [d for d in deposits if d.date.year == year and d.date.month == month]
        month_total = sum(d.amount for d in month_deposits)
        month_avg = month_total / len(month_deposits) if month_deposits else 0

        cells = [
            f"{month_name[month]} {year}",
            len(month_deposits),
            month_total,
            month_avg
        ]
        for col, value in enumerate(cells, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

    # User Balances Sheet
    wb.create_sheet("User Balances")
    ws_balances = wb["User Balances"]

    balance_headers = ["User Name", "Total Deposits", "Personal Withdrawals",
                       "Group Withdrawals", "Total Withdrawals", "Current Balance"]

    for col, header in enumerate(balance_headers, start=1):
        cell = ws_balances.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row, user in enumerate(users, start=2):
        total_deposits = user.total_deposited()
        personal_withdrawals = sum(w.amount for w in user.withdrawals.filter_by(withdrawal_type='personal'))
        group_withdrawals = sum(w.amount for w in user.withdrawals.filter_by(withdrawal_type='group'))
        total_withdrawals = personal_withdrawals + group_withdrawals
        balance = total_deposits - total_withdrawals

        cells = [
            user.name,
            total_deposits,
            personal_withdrawals,
            group_withdrawals,
            total_withdrawals,
            balance
        ]
        for col, value in enumerate(cells, start=1):
            cell = ws_balances.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

    # Monthly sheets
    for year, month in months:
        ws = wb.create_sheet(title=f"{month_name[month]} {year}")

        headers = ["Date", "User", "Amount"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        month_deposits = [d for d in deposits if d.date.year == year and d.date.month == month]
        for row, deposit in enumerate(month_deposits, start=2):
            cells = [deposit.date, deposit.user.name, deposit.amount]
            for col, value in enumerate(cells, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

        # Monthly total
        total_row = len(month_deposits) + 2
        ws.cell(row=total_row, column=2, value="Total").font = Font(bold=True)
        total_cell = ws.cell(row=total_row, column=3, value=f"=SUM(C2:C{total_row - 1})")
        total_cell.font = Font(bold=True)
        total_cell.border = border

    # Auto-adjust column widths
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    # Save to BytesIO object
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Generate timestamp for filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"deposit_report_{timestamp}.xlsx"

    return send_file(
        excel_file,
        download_name=filename,
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


from datetime import datetime
from dateutil.relativedelta import relativedelta
from sqlalchemy import func

from flask import send_file
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak, Frame, \
    NextPageTemplate, PageTemplate
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.graphics.shapes import Drawing, Rect
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.legends import Legend
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from io import BytesIO
import random
from datetime import datetime

# Try to register custom fonts, use default if not available
try:
    pdfmetrics.registerFont(TTFont('Roboto', 'Roboto-Regular.ttf'))
    pdfmetrics.registerFont(TTFont('Roboto-Bold', 'Roboto-Bold.ttf'))
    base_font = 'Roboto'
    bold_font = 'Roboto-Bold'
except:
    print("Roboto fonts not found. Using default fonts.")
    base_font = 'Helvetica'
    bold_font = 'Helvetica-Bold'


def add_pie_chart(data, labels):
    drawing = Drawing(6 * inch, 3 * inch)
    pie = Pie()
    pie.x = 150
    pie.y = 15
    pie.width = 150
    pie.height = 150
    pie.data = data
    pie.labels = None
    pie.slices.strokeWidth = 0.5
    pie.slices[0].fillColor = colors.HexColor("#3182CE")
    pie.slices[1].fillColor = colors.HexColor("#38A169")
    pie.slices[2].fillColor = colors.HexColor("#DD6B20")

    legend = Legend()
    legend.x = 330
    legend.y = 15
    legend.dx = 8
    legend.dy = 8
    legend.fontName = base_font
    legend.fontSize = 10
    legend.boxAnchor = "w"
    legend.columnMaximum = 1
    legend.strokeWidth = 0.5
    legend.strokeColor = colors.black
    legend.deltax = 75
    legend.deltay = 10
    legend.autoXPadding = 5
    legend.yGap = 0
    legend.dxTextSpace = 5
    legend.alignment = "right"
    legend.dividerLines = 1 | 2 | 4
    legend.dividerOffsY = 4.5
    legend.subCols.rpad = 30

    for i, color in enumerate([colors.HexColor("#3182CE"), colors.HexColor("#38A169"), colors.HexColor("#DD6B20")]):
        legend.colorNamePairs.append((color, labels[i]))

    drawing.add(pie)
    drawing.add(legend)
    return drawing


def add_line_chart(data, labels):
    drawing = Drawing(6 * inch, 3 * inch)
    lc = HorizontalLineChart()
    lc.x = 50
    lc.y = 50
    lc.height = 125
    lc.width = 350
    lc.data = data
    lc.categoryAxis.categoryNames = labels
    lc.categoryAxis.labels.boxAnchor = 'n'
    lc.categoryAxis.labels.angle = 0
    lc.categoryAxis.labels.dy = -10
    lc.categoryAxis.labels.fontName = base_font
    lc.valueAxis.labels.fontName = base_font
    lc.valueAxis.valueMin = 0
    lc.valueAxis.valueMax = max(max(d) for d in data) * 1.1
    lc.valueAxis.valueStep = max(max(d) for d in data) / 5
    lc.lines[0].strokeColor = colors.HexColor("#3182CE")
    lc.lines[1].strokeColor = colors.HexColor("#38A169")
    lc.joinedLines = 1
    drawing.add(lc)
    return drawing


def header_footer(canvas, doc):
    canvas.saveState()
    # Header
    canvas.setFont(bold_font, 10)
    canvas.drawString(0.5 * inch, 10.5 * inch, "Financial Report")
    canvas.drawRightString(7.5 * inch, 10.5 * inch, "Generated on: " + datetime.now().strftime("%Y-%m-%d"))
    # Footer
    canvas.setFont(base_font, 8)
    canvas.drawString(0.5 * inch, 0.5 * inch, "Confidential")
    canvas.drawRightString(7.5 * inch, 0.5 * inch, "Page %d" % doc.page)
    canvas.restoreState()


@app.route('/download_user_report/<int:user_id>')
@login_required
@admin_required
def download_user_report(user_id):
    # Get user data
    user = User.query.get_or_404(user_id)
    deposits = user.deposits.order_by(Deposit.date).all()
    personal_withdrawals = user.withdrawals.filter_by(withdrawal_type='personal').order_by(Withdrawal.date).all()
    group_withdrawals = user.withdrawals.filter_by(withdrawal_type='group').order_by(Withdrawal.date).all()

    # Generate filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{user.name}_report_{timestamp}.pdf"

    # Create PDF buffer
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=inch / 2, leftMargin=inch / 2,
                            topMargin=inch, bottomMargin=inch)

    styles = getSampleStyleSheet()
    elements = []

    # Define improved styles
    title_style = ParagraphStyle('Title', parent=styles['Heading1'],
                                 fontName=bold_font, fontSize=18, alignment=1, spaceAfter=24)
    heading_style = ParagraphStyle('Heading', parent=styles['Heading2'],
                                   fontName=bold_font, fontSize=14, spaceAfter=12)
    subheading_style = ParagraphStyle('SubHeading', parent=styles['Heading3'],
                                      fontName=bold_font, fontSize=12, spaceAfter=8)
    year_style = ParagraphStyle('YearHeading', parent=styles['Heading3'],
                                fontName=bold_font, fontSize=13, spaceAfter=10,
                                alignment=0, backColor=colors.lightgrey,
                                borderWidth=1, borderColor=colors.grey,
                                borderPadding=5)
    normal_style = ParagraphStyle('Normal', parent=styles['Normal'],
                                  fontName=base_font, fontSize=10, spaceAfter=6)
    cell_style = ParagraphStyle('CellText', parent=styles['Normal'],
                                fontName=base_font, fontSize=10, spaceAfter=3,
                                wordWrap='CJK', leading=12)

    # Add logo and title
    elements.append(Paragraph(f"Financial Report: {user.name}", title_style))
    elements.append(Spacer(1, 0.25 * inch))

    # Date of report generation
    elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y at %H:%M')}", normal_style))
    elements.append(Spacer(1, 0.5 * inch))

    # User Summary with improved formatting
    elements.append(Paragraph("User Summary", heading_style))

    summary_data = [
        ["Name:", user.name],
        ["Email:", user.email],
        ["Role:", user.role],
        ["Total Deposited:", f"{user.total_deposited():,.0f} Ksh"],
        ["Total Personal Withdrawals:", f"{sum(w.amount for w in personal_withdrawals):,.0f} Ksh"],
        ["Total Group Withdrawals:", f"{sum(w.amount for w in group_withdrawals):,.0f} Ksh"],
        ["Current Balance:", f"{user.balance():,.0f} Ksh"]
    ]

    summary_table = Table(summary_data, colWidths=[2.5 * inch, 4 * inch])
    summary_table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, -1), base_font, 10),
        ('FONT', (0, 0), (0, -1), bold_font, 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
        ('LINEBELOW', (0, -1), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey)
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.75 * inch))

    # Function to format date in a more readable way
    def format_date(date):
        return date.strftime("%d %b %Y")  # e.g., "17 Jul 2016"

    # Function to wrap text in Paragraph for table cells
    def wrap_text(text):
        if text:
            return Paragraph(text, cell_style)
        return ""

    # Group all transactions by year
    all_transactions = []
    for deposit in deposits:
        all_transactions.append({"type": "deposit", "date": deposit.date, "amount": deposit.amount, "reason": ""})

    for withdrawal in personal_withdrawals:
        all_transactions.append(
            {"type": "personal", "date": withdrawal.date, "amount": withdrawal.amount, "reason": withdrawal.reason})

    for withdrawal in group_withdrawals:
        all_transactions.append(
            {"type": "group", "date": withdrawal.date, "amount": withdrawal.amount, "reason": withdrawal.reason})

    # Sort all transactions by date
    all_transactions.sort(key=lambda x: x["date"])

    # Get unique years
    years = sorted(set(transaction["date"].year for transaction in all_transactions))

    if not years:
        elements.append(Paragraph("No transaction records found.", normal_style))
    else:
        # DEPOSIT SECTION
        elements.append(Paragraph("Deposit History", heading_style))

        if not deposits:
            elements.append(Paragraph("No deposit records found.", normal_style))
        else:
            by_year_deposits = {}
            for year in years:
                by_year_deposits[year] = [d for d in deposits if d.date.year == year]

            # Track if we're on the first year to avoid extra page break
            first_year = True

            for year in years:
                year_deposits = by_year_deposits.get(year, [])
                if year_deposits:
                    # Add a page break before each year's data except the first one
                    if not first_year:
                        elements.append(PageBreak())
                    first_year = False

                    elements.append(Paragraph(f"{year}", year_style))

                    deposit_data = [["Date", "Amount (Ksh)", "Running Balance"]]
                    running_balance = sum(d.amount for d in deposits if d.date.year < year)

                    for deposit in year_deposits:
                        running_balance += deposit.amount
                        deposit_data.append([
                            format_date(deposit.date),
                            f"{deposit.amount:,.0f}",
                            f"{running_balance:,.0f}"
                        ])

                    # Add a summary row for the year
                    deposit_data.append([
                        f"Total for {year}",
                        f"{sum(d.amount for d in year_deposits):,.0f}",
                        ""
                    ])

                    deposit_table = Table(deposit_data, colWidths=[2 * inch, 2 * inch, 2 * inch])
                    deposit_table.setStyle(TableStyle([
                        ('FONT', (0, 0), (-1, 0), bold_font, 10),
                        ('FONT', (0, 1), (-1, -1), base_font, 10),
                        ('FONT', (0, -1), (-1, -1), bold_font, 10),  # Bold for total row
                        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),  # Highlight total row
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                        ('ALIGN', (1, 0), (2, -1), 'RIGHT'),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                    ]))
                    elements.append(deposit_table)
                    elements.append(Spacer(1, 0.3 * inch))

        elements.append(PageBreak())

        # PERSONAL WITHDRAWAL SECTION
        elements.append(Paragraph("Personal Withdrawal History", heading_style))

        if not personal_withdrawals:
            elements.append(Paragraph("No personal withdrawal records found.", normal_style))
        else:
            by_year_personal = {}
            for year in years:
                by_year_personal[year] = [w for w in personal_withdrawals if w.date.year == year]

            # Reset first year flag for this section
            first_year = True

            for year in years:
                year_withdrawals = by_year_personal.get(year, [])
                if year_withdrawals:
                    # Add a page break before each year's data except the first one
                    if not first_year:
                        elements.append(PageBreak())
                    first_year = False

                    elements.append(Paragraph(f"{year}", year_style))

                    p_withdrawal_data = [["Date", "Amount (Ksh)", "Reason"]]
                    for withdrawal in year_withdrawals:
                        p_withdrawal_data.append([
                            format_date(withdrawal.date),
                            f"{withdrawal.amount:,.0f}",
                            wrap_text(withdrawal.reason)  # Wrap reason text for personal withdrawals
                        ])

                    # Add a summary row for the year
                    p_withdrawal_data.append([
                        f"Total for {year}",
                        f"{sum(w.amount for w in year_withdrawals):,.0f}",
                        ""
                    ])

                    p_withdrawal_table = Table(p_withdrawal_data, colWidths=[1.5 * inch, 1.5 * inch, 3 * inch])
                    p_withdrawal_table.setStyle(TableStyle([
                        ('FONT', (0, 0), (-1, 0), bold_font, 10),
                        ('FONT', (0, -1), (-1, -1), bold_font, 10),  # Bold for total row
                        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),  # Highlight total row
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                    ]))
                    elements.append(p_withdrawal_table)
                    elements.append(Spacer(1, 0.3 * inch))

        elements.append(PageBreak())

        # GROUP WITHDRAWAL SECTION - With improved text wrapping for reasons
        elements.append(Paragraph("Group Withdrawal History", heading_style))

        if not group_withdrawals:
            elements.append(Paragraph("No group withdrawal records found.", normal_style))
        else:
            by_year_group = {}
            for year in years:
                by_year_group[year] = [w for w in group_withdrawals if w.date.year == year]

            # Reset first year flag for this section
            first_year = True

            for year in years:
                year_withdrawals = by_year_group.get(year, [])
                if year_withdrawals:
                    # Add a page break before each year's data except the first one
                    if not first_year:
                        elements.append(PageBreak())
                    first_year = False

                    elements.append(Paragraph(f"{year}", year_style))

                    g_withdrawal_data = [["Date", "Amount (Ksh)", "Reason"]]
                    for withdrawal in year_withdrawals:
                        # Wrap the reason text in a Paragraph object for proper text wrapping
                        g_withdrawal_data.append([
                            format_date(withdrawal.date),
                            f"{withdrawal.amount:,.0f}",
                            wrap_text(withdrawal.reason)
                        ])

                    # Add a summary row for the year
                    g_withdrawal_data.append([
                        f"Total for {year}",
                        f"{sum(w.amount for w in year_withdrawals):,.0f}",
                        ""
                    ])

                    # Use automatic row heights to accommodate wrapped text
                    g_withdrawal_table = Table(g_withdrawal_data, colWidths=[1.5 * inch, 1.5 * inch, 3 * inch])
                    g_withdrawal_table.setStyle(TableStyle([
                        ('FONT', (0, 0), (-1, 0), bold_font, 10),
                        ('FONT', (0, -1), (-1, -1), bold_font, 10),  # Bold for total row
                        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),  # Highlight total row
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                    ]))
                    elements.append(g_withdrawal_table)
                    elements.append(Spacer(1, 0.3 * inch))

        elements.append(PageBreak())

        # YEARLY SUMMARY SECTION
        elements.append(Paragraph("Yearly Summary", heading_style))

        yearly_data = [["Year", "Deposits", "Withdrawals", "Net Flow"]]

        for year in years:
            year_deposits = sum(d.amount for d in deposits if d.date.year == year)

            year_personal = sum(w.amount for w in personal_withdrawals
                                if w.date.year == year)

            year_group = sum(w.amount for w in group_withdrawals
                             if w.date.year == year)

            total_withdrawals = year_personal + year_group
            net_flow = year_deposits - total_withdrawals

            yearly_data.append([
                f"{year}",
                f"{year_deposits:,.0f} Ksh",
                f"{total_withdrawals:,.0f} Ksh",
                f"{net_flow:,.0f} Ksh"
            ])

        # Add totals row
        total_deposits = sum(d.amount for d in deposits)
        total_withdrawals = sum(w.amount for w in personal_withdrawals + group_withdrawals)
        total_net = total_deposits - total_withdrawals

        yearly_data.append([
            "Grand Total",
            f"{total_deposits:,.0f} Ksh",
            f"{total_withdrawals:,.0f} Ksh",
            f"{total_net:,.0f} Ksh"
        ])

        yearly_table = Table(yearly_data, colWidths=[1.5 * inch, 1.5 * inch, 1.5 * inch, 1.5 * inch])
        yearly_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, 0), bold_font, 10),
            ('FONT', (0, 1), (-1, -1), base_font, 10),
            ('FONT', (0, -1), (-1, -1), bold_font, 10),  # Bold for total row
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),  # Highlight total row
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(yearly_table)

        # MONTHLY DETAIL SECTION (for the current year)
        current_year = datetime.now().year
        if current_year in years:
            elements.append(Spacer(1, 0.5 * inch))
            elements.append(Paragraph(f"Monthly Detail for {current_year}", heading_style))

            # Get monthly data for current year
            monthly_data = [["Month", "Deposits", "Withdrawals", "Net Flow"]]

            for month in range(1, 13):
                month_deposits = sum(d.amount for d in deposits
                                     if d.date.year == current_year and d.date.month == month)

                month_withdrawals = sum(w.amount for w in personal_withdrawals + group_withdrawals
                                        if w.date.year == current_year and w.date.month == month)

                # Only add months that have activity
                if month_deposits > 0 or month_withdrawals > 0:
                    net_flow = month_deposits - month_withdrawals

                    monthly_data.append([
                        f"{month_name[month]}",
                        f"{month_deposits:,.0f} Ksh",
                        f"{month_withdrawals:,.0f} Ksh",
                        f"{net_flow:,.0f} Ksh"
                    ])

            # Add year total row
            year_deposits = sum(d.amount for d in deposits if d.date.year == current_year)
            year_withdrawals = sum(w.amount for w in personal_withdrawals + group_withdrawals
                                   if w.date.year == current_year)
            year_net = year_deposits - year_withdrawals

            monthly_data.append([
                f"Total {current_year}",
                f"{year_deposits:,.0f} Ksh",
                f"{year_withdrawals:,.0f} Ksh",
                f"{year_net:,.0f} Ksh"
            ])

            # Only add the table if there are months with activity
            if len(monthly_data) > 2:  # Header + at least one month + total row
                monthly_table = Table(monthly_data, colWidths=[1.5 * inch, 1.5 * inch, 1.5 * inch, 1.5 * inch])
                monthly_table.setStyle(TableStyle([
                    ('FONT', (0, 0), (-1, 0), bold_font, 10),
                    ('FONT', (0, 1), (-1, -1), base_font, 10),
                    ('FONT', (0, -1), (-1, -1), bold_font, 10),  # Bold for total row
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),  # Highlight total row
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ]))
                elements.append(monthly_table)

    # Build PDF
    doc.build(elements, onFirstPage=header_footer, onLaterPages=header_footer)

    # Prepare response
    buffer.seek(0)
    return send_file(
        buffer,
        download_name=filename,
        as_attachment=True,
        mimetype="application/pdf"
    )


@app.route('/download_all_users_report')
@login_required
@admin_required
def download_all_users_report():
    # Get all users
    users = User.query.order_by(User.name).all()

    # Generate filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"all_users_report_{timestamp}.pdf"

    # Create PDF buffer
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=inch / 2, leftMargin=inch / 2,
                            topMargin=inch, bottomMargin=inch)

    styles = getSampleStyleSheet()
    elements = []

    # Define styles
    title_style = ParagraphStyle('Title', parent=styles['Heading1'],
                                 fontName=bold_font, fontSize=18, alignment=1, spaceAfter=24)
    heading_style = ParagraphStyle('Heading', parent=styles['Heading2'],
                                   fontName=bold_font, fontSize=14, spaceAfter=12)
    subheading_style = ParagraphStyle('SubHeading', parent=styles['Heading3'],
                                      fontName=bold_font, fontSize=12, spaceAfter=8)
    normal_style = ParagraphStyle('Normal', parent=styles['Normal'],
                                  fontName=base_font, fontSize=10, spaceAfter=6)
    cell_style = ParagraphStyle('CellText', parent=styles['Normal'],
                                fontName=base_font, fontSize=10, spaceAfter=3,
                                wordWrap='CJK', leading=12)

    # Add title and date of generation
    elements.append(Paragraph("All Users Financial Report", title_style))
    elements.append(Spacer(1, 0.25 * inch))
    elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y at %H:%M')}", normal_style))
    elements.append(Spacer(1, 0.5 * inch))

    # User Summary Section
    elements.append(Paragraph("Summary Statistics", heading_style))

    # Calculate summary statistics
    total_users = len(users)
    total_deposited = sum(user.total_deposited() for user in users)
    total_personal_withdrawals = sum(
        sum(w.amount for w in user.withdrawals.filter_by(withdrawal_type='personal').all())
        for user in users
    )
    total_group_withdrawals = sum(
        sum(w.amount for w in user.withdrawals.filter_by(withdrawal_type='group').all())
        for user in users
    )
    total_balance = sum(user.balance() for user in users)

    summary_data = [
        ["Total Users:", f"{total_users}"],
        ["Total Deposits:", f"{total_deposited:,.0f} Ksh"],
        ["Total Personal Withdrawals:", f"{total_personal_withdrawals:,.0f} Ksh"],
        ["Total Group Withdrawals:", f"{total_group_withdrawals:,.0f} Ksh"],
        ["Total Current Balance:", f"{total_balance:,.0f} Ksh"]
    ]

    summary_table = Table(summary_data, colWidths=[3 * inch, 3.5 * inch])
    summary_table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, -1), base_font, 10),
        ('FONT', (0, 0), (0, -1), bold_font, 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
        ('LINEBELOW', (0, -1), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey)
    ]))

    elements.append(summary_table)
    elements.append(Spacer(1, 0.5 * inch))

    # Users Detailed Overview
    elements.append(Paragraph("User Details", heading_style))

    user_details_data = [["Name", "Role", "Deposits", "Personal W/D", "Group W/D", "Balance"]]

    for user in users:
        deposits = user.total_deposited()
        personal_withdrawals = sum(w.amount for w in user.withdrawals.filter_by(withdrawal_type='personal').all())
        group_withdrawals = sum(w.amount for w in user.withdrawals.filter_by(withdrawal_type='group').all())
        balance = user.balance()

        user_details_data.append([
            user.name,
            user.role,
            f"{deposits:,.0f}",
            f"{personal_withdrawals:,.0f}",
            f"{group_withdrawals:,.0f}",
            f"{balance:,.0f}"
        ])

    # Calculate column totals for the table footer
    total_deposits = sum(user.total_deposited() for user in users)
    total_personal = sum(
        sum(w.amount for w in user.withdrawals.filter_by(withdrawal_type='personal').all())
        for user in users
    )
    total_group = sum(
        sum(w.amount for w in user.withdrawals.filter_by(withdrawal_type='group').all())
        for user in users
    )
    total_bal = sum(user.balance() for user in users)

    # Add totals row
    user_details_data.append([
        "TOTALS",
        "",
        f"{total_deposits:,.0f}",
        f"{total_personal:,.0f}",
        f"{total_group:,.0f}",
        f"{total_bal:,.0f}"
    ])

    # Create table with appropriate column widths
    user_details_table = Table(user_details_data,
                               colWidths=[1.5 * inch, 1 * inch, 1.25 * inch, 1.25 * inch, 1.25 * inch, 1.25 * inch])
    user_details_table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, 0), bold_font, 10),
        ('FONT', (0, 1), (-1, -2), base_font, 10),
        ('FONT', (0, -1), (-1, -1), bold_font, 10),  # Bold for total row
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))

    elements.append(user_details_table)
    elements.append(PageBreak())

    # Year-by-Year Analysis
    elements.append(Paragraph("Yearly Analysis", heading_style))

    # Get all transactions to find the range of years
    all_deposits = Deposit.query.order_by(Deposit.date).all()
    all_withdrawals = Withdrawal.query.order_by(Withdrawal.date).all()

    if not all_deposits and not all_withdrawals:
        elements.append(Paragraph("No transaction history found.", normal_style))
    else:
        # Extract all years from transactions
        all_years = set()
        for deposit in all_deposits:
            all_years.add(deposit.date.year)
        for withdrawal in all_withdrawals:
            all_years.add(withdrawal.date.year)

        years = sorted(all_years)

        # Create yearly analysis
        yearly_data = [["Year", "Deposits", "Withdrawals", "Net Flow"]]

        for year in years:
            # Calculate yearly totals
            year_deposits = sum(d.amount for d in all_deposits if d.date.year == year)
            year_withdrawals = sum(w.amount for w in all_withdrawals if w.date.year == year)
            net_flow = year_deposits - year_withdrawals

            yearly_data.append([
                f"{year}",
                f"{year_deposits:,.0f} Ksh",
                f"{year_withdrawals:,.0f} Ksh",
                f"{net_flow:,.0f} Ksh"
            ])

        # Add totals row
        total_deposits = sum(d.amount for d in all_deposits)
        total_withdrawals = sum(w.amount for w in all_withdrawals)
        total_net = total_deposits - total_withdrawals

        yearly_data.append([
            "Grand Total",
            f"{total_deposits:,.0f} Ksh",
            f"{total_withdrawals:,.0f} Ksh",
            f"{total_net:,.0f} Ksh"
        ])

        yearly_table = Table(yearly_data, colWidths=[1.5 * inch, 2 * inch, 2 * inch, 2 * inch])
        yearly_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, 0), bold_font, 10),
            ('FONT', (0, 1), (-1, -2), base_font, 10),
            ('FONT', (0, -1), (-1, -1), bold_font, 10),  # Bold for total row
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))

        elements.append(yearly_table)
        elements.append(Spacer(1, 0.5 * inch))

        # Current Year Monthly Analysis
        current_year = datetime.now().year
        if current_year in years:
            elements.append(Paragraph(f"Monthly Detail for {current_year}", subheading_style))

            # Calculate monthly data for current year
            monthly_data = [["Month", "Deposits", "Withdrawals", "Net Flow"]]

            for month in range(1, 13):
                month_deposits = sum(d.amount for d in all_deposits
                                     if d.date.year == current_year and d.date.month == month)

                month_withdrawals = sum(w.amount for w in all_withdrawals
                                        if w.date.year == current_year and w.date.month == month)

                # Only include months with activity
                if month_deposits > 0 or month_withdrawals > 0:
                    month_name = datetime(current_year, month, 1).strftime('%B')
                    net_flow = month_deposits - month_withdrawals

                    monthly_data.append([
                        month_name,
                        f"{month_deposits:,.0f} Ksh",
                        f"{month_withdrawals:,.0f} Ksh",
                        f"{net_flow:,.0f} Ksh"
                    ])

            # Add year total row
            year_deposits = sum(d.amount for d in all_deposits if d.date.year == current_year)
            year_withdrawals = sum(w.amount for w in all_withdrawals if w.date.year == current_year)
            year_net = year_deposits - year_withdrawals

            monthly_data.append([
                f"Total {current_year}",
                f"{year_deposits:,.0f} Ksh",
                f"{year_withdrawals:,.0f} Ksh",
                f"{year_net:,.0f} Ksh"
            ])

            # Only add the table if there are months with activity
            if len(monthly_data) > 2:  # Header + at least one month + total row
                monthly_table = Table(monthly_data, colWidths=[1.5 * inch, 2 * inch, 2 * inch, 2 * inch])
                monthly_table.setStyle(TableStyle([
                    ('FONT', (0, 0), (-1, 0), bold_font, 10),
                    ('FONT', (0, 1), (-1, -2), base_font, 10),
                    ('FONT', (0, -1), (-1, -1), bold_font, 10),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ]))
                elements.append(monthly_table)

    # Top Contributors Section
    elements.append(PageBreak())
    elements.append(Paragraph("Top Contributors", heading_style))

    # Sort users by total deposited amount (descending)
    top_depositors = sorted(users, key=lambda u: u.total_deposited(), reverse=True)[:10]  # Top 10

    if top_depositors:
        top_depositors_data = [["Rank", "Name", "Total Deposited"]]

        for i, user in enumerate(top_depositors, 1):
            top_depositors_data.append([
                f"{i}",
                user.name,
                f"{user.total_deposited():,.0f} Ksh"
            ])

        top_depositors_table = Table(top_depositors_data, colWidths=[1 * inch, 3 * inch, 3 * inch])
        top_depositors_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, 0), bold_font, 10),
            ('FONT', (0, 1), (-1, -1), base_font, 10),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))

        elements.append(top_depositors_table)
    else:
        elements.append(Paragraph("No deposit records found.", normal_style))

    # Build PDF with header/footer
    doc.build(elements, onFirstPage=header_footer, onLaterPages=header_footer)

    # Prepare response
    buffer.seek(0)
    return send_file(
        buffer,
        download_name=filename,
        as_attachment=True,
        mimetype="application/pdf"
    )


@app.route('/download_comprehensive_users_report')
@login_required
@admin_required
def download_comprehensive_users_report():
    # Get all users
    users = User.query.order_by(User.name).all()

    # Generate filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"comprehensive_users_report_{timestamp}.pdf"

    # Create PDF buffer
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=inch / 2, leftMargin=inch / 2,
                            topMargin=inch, bottomMargin=inch)

    styles = getSampleStyleSheet()
    elements = []

    # Define styles
    title_style = ParagraphStyle('Title', parent=styles['Heading1'],
                                 fontName=bold_font, fontSize=18, alignment=1, spaceAfter=24)
    heading_style = ParagraphStyle('Heading', parent=styles['Heading2'],
                                   fontName=bold_font, fontSize=14, spaceAfter=12)
    subheading_style = ParagraphStyle('SubHeading', parent=styles['Heading3'],
                                      fontName=bold_font, fontSize=12, spaceAfter=8)
    normal_style = ParagraphStyle('Normal', parent=styles['Normal'],
                                  fontName=base_font, fontSize=10, spaceAfter=6)
    cell_style = ParagraphStyle('CellText', parent=styles['Normal'],
                                fontName=base_font, fontSize=9, spaceAfter=3,
                                wordWrap='CJK', leading=11)

    # Format functions
    def format_currency(amount):
        return f"{amount:,.0f} Ksh"

    def wrap_text(text):
        if text:
            return Paragraph(text, cell_style)
        return ""

    # Add title and date of generation
    elements.append(Paragraph("Comprehensive Users Financial Report", title_style))
    elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y at %H:%M')}", normal_style))
    elements.append(Spacer(1, 0.5 * inch))

    # Organization Summary Section
    elements.append(Paragraph("Organization Summary", heading_style))

    # Calculate overall statistics
    total_users = len(users)
    total_deposited = sum(user.total_deposited() for user in users)
    total_personal_withdrawals = sum(
        sum(w.amount for w in user.withdrawals.filter_by(withdrawal_type='personal').all())
        for user in users
    )
    total_group_withdrawals = sum(
        sum(w.amount for w in user.withdrawals.filter_by(withdrawal_type='group').all())
        for user in users
    )
    total_withdrawals = total_personal_withdrawals + total_group_withdrawals
    total_balance = sum(user.balance() for user in users)

    # Get all transactions to find the range of years
    all_deposits = Deposit.query.order_by(Deposit.date).all()
    all_withdrawals = Withdrawal.query.order_by(Withdrawal.date).all()

    all_years = set()
    if all_deposits or all_withdrawals:
        for deposit in all_deposits:
            all_years.add(deposit.date.year)
        for withdrawal in all_withdrawals:
            all_years.add(withdrawal.date.year)

    years = sorted(all_years)

    # Create organization summary
    summary_data = [
        ["Total Users:", f"{total_users}"],
        ["Active Years:", f"{min(years, default='N/A')} - {max(years, default='N/A')}"],
        ["Total Deposits:", format_currency(total_deposited)],
        ["Total Personal Withdrawals:", format_currency(total_personal_withdrawals)],
        ["Total Group Withdrawals:", format_currency(total_group_withdrawals)],
        ["Total All Withdrawals:", format_currency(total_withdrawals)],
        ["Current Total Balance:", format_currency(total_balance)]
    ]

    summary_table = Table(summary_data, colWidths=[3 * inch, 3.5 * inch])
    summary_table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, -1), base_font, 10),
        ('FONT', (0, 0), (0, -1), bold_font, 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
        ('LINEBELOW', (0, -1), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey)
    ]))

    elements.append(summary_table)

    # ====================== YEARLY OVERVIEW SECTION ======================
    elements.append(PageBreak())
    elements.append(Paragraph("Yearly Financial Overview", heading_style))

    if not years:
        elements.append(Paragraph("No transaction history found.", normal_style))
    else:
        # Create yearly overview table
        yearly_data = [["Year", "Deposits", "Personal W/D", "Group W/D", "Total W/D", "Net Flow"]]

        for year in years:
            # Calculate yearly totals
            year_deposits = sum(d.amount for d in all_deposits if d.date.year == year)
            year_personal = sum(w.amount for w in all_withdrawals
                                if w.date.year == year and w.withdrawal_type == 'personal')
            year_group = sum(w.amount for w in all_withdrawals
                             if w.date.year == year and w.withdrawal_type == 'group')
            year_withdrawals = year_personal + year_group
            net_flow = year_deposits - year_withdrawals

            yearly_data.append([
                f"{year}",
                format_currency(year_deposits),
                format_currency(year_personal),
                format_currency(year_group),
                format_currency(year_withdrawals),
                format_currency(net_flow)
            ])

        # Add totals row
        total_deposits = sum(d.amount for d in all_deposits)
        total_personal = sum(w.amount for w in all_withdrawals if w.withdrawal_type == 'personal')
        total_group = sum(w.amount for w in all_withdrawals if w.withdrawal_type == 'group')
        total_all_withdrawals = total_personal + total_group
        total_net = total_deposits - total_all_withdrawals

        yearly_data.append([
            "Grand Total",
            format_currency(total_deposits),
            format_currency(total_personal),
            format_currency(total_group),
            format_currency(total_all_withdrawals),
            format_currency(total_net)
        ])

        yearly_table = Table(yearly_data,
                             colWidths=[1.3 * inch, 1.8 * inch, 1.8 * inch, 1.8 * inch, 1.8 * inch, 1.8 * inch])
        yearly_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, 0), bold_font, 10),
            ('FONT', (0, 1), (-1, -2), base_font, 10),
            ('FONT', (0, -1), (-1, -1), bold_font, 10),  # Bold for total row
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))

        elements.append(yearly_table)

    # ====================== MONTHLY OVERVIEW FOR CURRENT YEAR ======================
    current_year = datetime.now().year
    if current_year in years:
        elements.append(Spacer(1, 0.5 * inch))
        elements.append(Paragraph(f"Monthly Overview for {current_year}", subheading_style))

        # Create monthly overview table
        monthly_data = [["Month", "Deposits", "Personal W/D", "Group W/D", "Total W/D", "Net Flow"]]

        for month in range(1, 13):
            month_name = datetime(current_year, month, 1).strftime('%B')

            month_deposits = sum(d.amount for d in all_deposits
                                 if d.date.year == current_year and d.date.month == month)

            month_personal = sum(w.amount for w in all_withdrawals
                                 if w.date.year == current_year and w.date.month == month
                                 and w.withdrawal_type == 'personal')

            month_group = sum(w.amount for w in all_withdrawals
                              if w.date.year == current_year and w.date.month == month
                              and w.withdrawal_type == 'group')

            month_withdrawals = month_personal + month_group
            net_flow = month_deposits - month_withdrawals

            # Only include months with activity
            if month_deposits > 0 or month_withdrawals > 0:
                monthly_data.append([
                    month_name,
                    format_currency(month_deposits),
                    format_currency(month_personal),
                    format_currency(month_group),
                    format_currency(month_withdrawals),
                    format_currency(net_flow)
                ])

        # Add year total row
        year_deposits = sum(d.amount for d in all_deposits if d.date.year == current_year)
        year_personal = sum(w.amount for w in all_withdrawals
                            if w.date.year == current_year and w.withdrawal_type == 'personal')
        year_group = sum(w.amount for w in all_withdrawals
                         if w.date.year == current_year and w.withdrawal_type == 'group')
        year_withdrawals = year_personal + year_group
        year_net = year_deposits - year_withdrawals

        monthly_data.append([
            f"Total {current_year}",
            format_currency(year_deposits),
            format_currency(year_personal),
            format_currency(year_group),
            format_currency(year_withdrawals),
            format_currency(year_net)
        ])

        # Only add the table if there are months with activity
        if len(monthly_data) > 2:  # Header + at least one month + total row
            monthly_table = Table(monthly_data,
                                  colWidths=[1.3 * inch, 1.8 * inch, 1.8 * inch, 1.8 * inch, 1.8 * inch, 1.8 * inch])
            monthly_table.setStyle(TableStyle([
                ('FONT', (0, 0), (-1, 0), bold_font, 10),
                ('FONT', (0, 1), (-1, -2), base_font, 10),
                ('FONT', (0, -1), (-1, -1), bold_font, 10),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ]))
            elements.append(monthly_table)

    # ====================== USER BY USER ANALYSIS ======================
    for user in users:
        elements.append(PageBreak())
        elements.append(Paragraph(f"User Report: {user.name}", heading_style))

        # Get user transactions
        deposits = user.deposits.order_by(Deposit.date).all()
        personal_withdrawals = user.withdrawals.filter_by(withdrawal_type='personal').order_by(Withdrawal.date).all()
        group_withdrawals = user.withdrawals.filter_by(withdrawal_type='group').order_by(Withdrawal.date).all()

        # User Summary
        user_summary_data = [
            ["Email:", user.email],
            ["Role:", user.role],
            ["Total Deposited:", format_currency(user.total_deposited())],
            ["Total Personal Withdrawals:", format_currency(sum(w.amount for w in personal_withdrawals))],
            ["Total Group Withdrawals:", format_currency(sum(w.amount for w in group_withdrawals))],
            ["Current Balance:", format_currency(user.balance())]
        ]

        user_summary_table = Table(user_summary_data, colWidths=[3 * inch, 6 * inch])
        user_summary_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), base_font, 10),
            ('FONT', (0, 0), (0, -1), bold_font, 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
            ('LINEBELOW', (0, -1), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey)
        ]))
        elements.append(user_summary_table)
        elements.append(Spacer(1, 0.3 * inch))

        # Create yearly breakdown for this user
        user_years = set()
        for deposit in deposits:
            user_years.add(deposit.date.year)
        for withdrawal in personal_withdrawals + group_withdrawals:
            user_years.add(withdrawal.date.year)

        user_years = sorted(user_years)

        if not user_years:
            elements.append(Paragraph("No transaction history found for this user.", normal_style))
            continue

        # Yearly breakdown table
        elements.append(Paragraph("Yearly Breakdown", subheading_style))

        yearly_user_data = [["Year", "Deposits", "Personal W/D", "Group W/D", "Total W/D", "Net Flow"]]

        for year in user_years:
            # Calculate yearly totals for this user
            year_deposits = sum(d.amount for d in deposits if d.date.year == year)
            year_personal = sum(w.amount for w in personal_withdrawals if w.date.year == year)
            year_group = sum(w.amount for w in group_withdrawals if w.date.year == year)
            year_withdrawals = year_personal + year_group
            net_flow = year_deposits - year_withdrawals

            yearly_user_data.append([
                f"{year}",
                format_currency(year_deposits),
                format_currency(year_personal),
                format_currency(year_group),
                format_currency(year_withdrawals),
                format_currency(net_flow)
            ])

        # Add totals row
        user_total_deposits = sum(d.amount for d in deposits)
        user_total_personal = sum(w.amount for w in personal_withdrawals)
        user_total_group = sum(w.amount for w in group_withdrawals)
        user_total_withdrawals = user_total_personal + user_total_group
        user_total_net = user_total_deposits - user_total_withdrawals

        yearly_user_data.append([
            "Total",
            format_currency(user_total_deposits),
            format_currency(user_total_personal),
            format_currency(user_total_group),
            format_currency(user_total_withdrawals),
            format_currency(user_total_net)
        ])

        yearly_user_table = Table(yearly_user_data,
                                  colWidths=[1.3 * inch, 1.8 * inch, 1.8 * inch, 1.8 * inch, 1.8 * inch, 1.8 * inch])
        yearly_user_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, 0), bold_font, 10),
            ('FONT', (0, 1), (-1, -2), base_font, 10),
            ('FONT', (0, -1), (-1, -1), bold_font, 10),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(yearly_user_table)

        # Show monthly breakdown for current year if user has activity
        current_year = datetime.now().year
        if current_year in user_years:
            elements.append(Spacer(1, 0.3 * inch))
            elements.append(Paragraph(f"Monthly Breakdown for {current_year}", subheading_style))

            monthly_user_data = [["Month", "Deposits", "Personal W/D", "Group W/D", "Total W/D", "Net Flow"]]
            has_activity = False

            for month in range(1, 13):
                month_name = datetime(current_year, month, 1).strftime('%B')

                month_deposits = sum(d.amount for d in deposits
                                     if d.date.year == current_year and d.date.month == month)

                month_personal = sum(w.amount for w in personal_withdrawals
                                     if w.date.year == current_year and w.date.month == month)

                month_group = sum(w.amount for w in group_withdrawals
                                  if w.date.year == current_year and w.date.month == month)

                month_withdrawals = month_personal + month_group
                net_flow = month_deposits - month_withdrawals

                # Only include months with activity
                if month_deposits > 0 or month_withdrawals > 0:
                    has_activity = True
                    monthly_user_data.append([
                        month_name,
                        format_currency(month_deposits),
                        format_currency(month_personal),
                        format_currency(month_group),
                        format_currency(month_withdrawals),
                        format_currency(net_flow)
                    ])

            # Add year total row
            year_deposits = sum(d.amount for d in deposits if d.date.year == current_year)
            year_personal = sum(w.amount for w in personal_withdrawals if w.date.year == current_year)
            year_group = sum(w.amount for w in group_withdrawals if w.date.year == current_year)
            year_withdrawals = year_personal + year_group
            year_net = year_deposits - year_withdrawals

            if has_activity:
                monthly_user_data.append([
                    f"Total {current_year}",
                    format_currency(year_deposits),
                    format_currency(year_personal),
                    format_currency(year_group),
                    format_currency(year_withdrawals),
                    format_currency(year_net)
                ])

                monthly_user_table = Table(monthly_user_data,
                                           colWidths=[1.3 * inch, 1.8 * inch, 1.8 * inch, 1.8 * inch, 1.8 * inch,
                                                      1.8 * inch])
                monthly_user_table.setStyle(TableStyle([
                    ('FONT', (0, 0), (-1, 0), bold_font, 10),
                    ('FONT', (0, 1), (-1, -2), base_font, 10),
                    ('FONT', (0, -1), (-1, -1), bold_font, 10),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ]))
                elements.append(monthly_user_table)
            else:
                elements.append(Paragraph(f"No activity recorded for {current_year}.", normal_style))

        # If the user has withdrawals with reasons, show a section of recent activity with reasons
        recent_withdrawals = (user.withdrawals.order_by(Withdrawal.date.desc())
                              .limit(5).all())

        if recent_withdrawals:
            elements.append(Spacer(1, 0.3 * inch))
            elements.append(Paragraph("Recent Withdrawal Activity", subheading_style))

            withdrawal_data = [["Date", "Type", "Amount", "Reason"]]
            for withdrawal in recent_withdrawals:
                withdrawal_data.append([
                    withdrawal.date.strftime("%d %b %Y"),
                    withdrawal.withdrawal_type.capitalize(),
                    format_currency(withdrawal.amount),
                    wrap_text(withdrawal.reason)
                ])

            withdrawal_table = Table(withdrawal_data, colWidths=[1.5 * inch, 1.5 * inch, 1.5 * inch, 5 * inch])
            withdrawal_table.setStyle(TableStyle([
                ('FONT', (0, 0), (-1, 0), bold_font, 10),
                ('FONT', (0, 1), (-1, -1), base_font, 10),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ]))
            elements.append(withdrawal_table)

    # ====================== TOP USERS SECTION ======================
    elements.append(PageBreak())
    elements.append(Paragraph("Top Contributors and Account Holders", heading_style))

    # Top depositors
    elements.append(Paragraph("Top 10 Contributors (By Total Deposits)", subheading_style))
    top_depositors = sorted(users, key=lambda u: u.total_deposited(), reverse=True)[:10]

    if top_depositors:
        top_depositors_data = [["Rank", "Name", "Role", "Total Deposited", "% of All Deposits"]]

        for i, user in enumerate(top_depositors, 1):
            user_deposit = user.total_deposited()
            percentage = (user_deposit / total_deposited * 100) if total_deposited > 0 else 0

            top_depositors_data.append([
                f"{i}",
                user.name,
                user.role,
                format_currency(user_deposit),
                f"{percentage:.1f}%"
            ])

        top_depositors_table = Table(top_depositors_data,
                                     colWidths=[0.7 * inch, 3 * inch, 1.5 * inch, 2 * inch, 2 * inch])
        top_depositors_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, 0), bold_font, 10),
            ('FONT', (0, 1), (-1, -1), base_font, 10),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (3, 0), (4, -1), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(top_depositors_table)
    else:
        elements.append(Paragraph("No deposit records found.", normal_style))

    # Top account holders (by balance)
    elements.append(Spacer(1, 0.5 * inch))
    elements.append(Paragraph("Top 10 Account Holders (By Current Balance)", subheading_style))
    top_balances = sorted(users, key=lambda u: u.balance(), reverse=True)[:10]

    if top_balances:
        top_balances_data = [["Rank", "Name", "Role", "Current Balance", "% of Total Balance"]]

        for i, user in enumerate(top_balances, 1):
            user_balance = user.balance()
            percentage = (user_balance / total_balance * 100) if total_balance > 0 else 0

            top_balances_data.append([
                f"{i}",
                user.name,
                user.role,
                format_currency(user_balance),
                f"{percentage:.1f}%"
            ])

        top_balances_table = Table(top_balances_data,
                                   colWidths=[0.7 * inch, 3 * inch, 1.5 * inch, 2 * inch, 2 * inch])
        top_balances_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, 0), bold_font, 10),
            ('FONT', (0, 1), (-1, -1), base_font, 10),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (3, 0), (4, -1), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(top_balances_table)
    else:
        elements.append(Paragraph("No balance records found.", normal_style))

    # Build PDF
    doc.build(elements, onFirstPage=header_footer, onLaterPages=header_footer)

    # Prepare response
    buffer.seek(0)
    return send_file(
        buffer,
        download_name=filename,
        as_attachment=True,
        mimetype="application/pdf"
    )

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True)